import openpyxl as op


def getData (entity, searchAttribute = None, searchString = None):
    workbook = op.load_workbook('./dao/db.xlsx')
    db = workbook[entity]

    if searchString == None or searchAttribute == None:
       rows_iter = db.iter_rows(min_col = 1, min_row = 2, max_col = db.max_column, max_row = db.max_row)
       valIter = [[cell.value for cell in list(row)] for row in rows_iter]
       return valIter
    else:
       valIter = []
       for row in db.iter_rows(values_only=True):        
            if row[searchAttribute] == searchString:            
                valIter.append(row)                
       return valIter  

        
def setData (entity, searchAttribute, searchString, dataRow, data):
   workbook = op.load_workbook('./dao/db.xlsx')
   db = workbook[entity]
   
   for row in db.iter_rows(values_only=True):
            i=1        
            if row[searchAttribute] == searchString:
                db.cell(row=i, column=5, value = data)
            else:
                i=i+1   